import os
from openpyxl import load_workbook
from fake_useragent import UserAgent
import requests
import json

wb = load_workbook('source.xlsx')

sheet = wb['Лист1']

count = 0

os.chdir('./data_for_crawling')

headers = {
    'User-Agent': UserAgent().random
}

for row in range(239, 240):

    try:

        test_link = sheet.cell(row=row, column=7).value

        source_json = dict()

        if sheet.cell(row=row, column=4).value.endswith('.'):
            source_json['name'] = sheet.cell(row=row, column=4).value[:-1]
        else:
            source_json['name'] = sheet.cell(row=row, column=4).value

        if sheet.cell(row=row, column=5).value.endswith('/'):
            source_json['link'] = sheet.cell(row=row, column=5).value[:-1]
        else:
            source_json['link'] = sheet.cell(row=row, column=5).value

        if sheet.cell(row=row, column=7).value.endswith('/'):
            start_link = sheet.cell(row=row, column=7).value[:-1]
        else:
            start_link = sheet.cell(row=row, column=7).value

        if sheet.cell(row=row, column=6).value.strip() == 'xml':
            type_link = 'rss'
        elif sheet.cell(row=row, column=6).value.strip() == 'robots.txt':
            type_link = 'robots.txt'
        else:
            type_link = sheet.cell(row=row, column=6).value.strip()

        try:
            # print(f'TEST LINK: {test_link}')
            # print(sheet.cell(row=row, column=6).value.strip())
            res = requests.get(test_link, headers=headers, verify=False)
            res.raise_for_status()
        except Exception as exp:
            print(f'FAILED LINK:{test_link}, SOURCE:{sheet.cell(row=row, column=4).value[:-1]}, {exp}')
            continue

        source_json['source_meta'] = {}
        source_json['source_meta']['type'] = sheet.cell(row=row, column=8).value
        source_json['source_meta']['license'] = sheet.cell(row=row, column=1).value
        source_json['source_meta']['founders'] = sheet.cell(row=row, column=10).value
        source_json['source_meta']['address'] = sheet.cell(row=row, column=11).value
        source_json['source_meta']['language'] = sheet.cell(row=row, column=12).value
        source_json['crawl_data'] = {}
        source_json['crawl_data']['update_period'] = 300
        source_json['crawl_data']['is_debug'] = "False"
        source_json['crawl_data']['requests_data'] = {
            "headers": {},
            "start_link": start_link,
            "type": type_link,
            "allow_redirects": "True",
            "index": []
        }
        source_json['crawl_data']['main_link'] = source_json['link']
        source_json['crawl_data']['links_tags'] = []
        source_json['crawl_data']['title_tag'] = ""
        source_json['crawl_data']['author_tags'] = ""
        source_json['crawl_data']['body_tag'] = []
        source_json['crawl_data']['decompose_tags'] = []
        source_json['crawl_data']['text_tags'] = ""

        os.mkdir(source_json['name'])
        with open(f'{source_json["name"]}/{source_json["name"]}.json', 'w') as file:
            json.dump(source_json, file, ensure_ascii=False, indent=4)

    except:
        print(f'{sheet.cell(row=row, column=4).value} NOT ADDED!')

    count += 1

