from flask import Flask, request, render_template

app = Flask(__name__)

@app.route('/')
def my_form():
    return render_template('form.html')

@app.route('/', methods=['POST'])
def my_form_post():
    
    #convert to lowercase
    text_in = request.form['text1'].lower()
    text = text_in
    import operator
    import re
    text = re.sub("[a-zA-Z0-9]", "", text)
    text = re.sub("[a-z]", "", text)
    import string
    print(string.punctuation)
    spec_chars = string.punctuation + '\n\xa0«»\t—…[]~' 
    text = "".join([ch for ch in text if ch not in spec_chars])
    def remove_chars_from_text(text, chars):
        return "".join([ch for ch in text if ch not in chars])
    text = remove_chars_from_text(text, spec_chars)
    text = remove_chars_from_text(text, string.digits)
    import nltk
    nltk.download('punkt')
    from nltk import word_tokenize
    text_tokens = word_tokenize(text)
    print (len(text_tokens))
    g = 0
    text_tokens_int = []
    while g < len(text_tokens):
        word = text_tokens[g]
        text_tokens_int.append(word)
        g += 1
    ### HERE START PSYCHOSEMANTIC CODE
    text_w = text_tokens_int
    r = 0
    full_list = []
    St_list = []
    St_list_all = []
    while r < len(text_w):
        word = text_w[r]
        print (word)
        word = word.lower()
        print (word)
        from heapq import nlargest
        import openpyxl
        wb = openpyxl.reader.excel.load_workbook(filename="psy_dataset.xlsx")
        print (wb.sheetnames)
        wb.active = 0
        sheet = wb.active
        print (sheet['A26'].value)
        print (sheet['A26'].value)
        print (sheet.cell(2,1).value)
        i = 0
        n = 1
        word_l = list(word.strip(" "))
        print (word_l)
        print (len(word_l))
        soft_let = 'е ё и ь ю я'.split()
        soft_let_pos = 'б в г д з к л м н п р с т ф х'.split()
        print (soft_let)
        print (soft_let_pos)
        result = list(set(word_l) - set(soft_let_pos))
        print (result, 'сравнение двух списков')
        slpn = 0
        sln = 0
        print (word_l)
        print (len(word_l))
        print (len(soft_let))
        for i in range(len(word_l)):
            slpn = 0
            while slpn < len(soft_let_pos):
                if word_l[i] == soft_let_pos[slpn]:
                    try:
                        print (word_l[i],' вот эта буква')
                        sln = 0
                        while sln < len(soft_let):
                            if word_l[i+1] == soft_let[sln]:               
                                try:
                                    print (word_l[i+1])
                                    word_l[i] = word_l[i] + '*'
                                    sln += 1
                                    slpn += 1
                                except IndexError:
                                    #print ('nah')
                                    sln += 1
                                    slpn += 1
                            else:
                                #print ('no')
                                sln += 1
                                slpn += 1
                    except IndexError:
                        print (word_l[i],' вот эта буква')
                        slpn += 1
                else:
                    slpn += 1
        print (word_l)
        i = 0
        n = 1
        cr = 1 #row
        cc = 1 #column

        #####   НОВЫЙ КОД    #####
        cck = 2
        print (sheet.cell(cr,cc).value, 'THIS IS TEST')


        word_lgood = []
        word_freq_uno = []
        word_char = []
        word_freq = []
        #print (word_lgood,'word_lgood')
        #print ('cck')
        while cck < 27:
            word_ldub = word_l[:]
            while i < len(word_l):
                #print('yeah')
                while True:
                    word_l[i] != sheet.cell(cr,cc).value
                    cr += 1
                    if word_l[i] == sheet.cell(cr,cc).value:
                        word_l[i] = sheet.cell(cr,cc+1).value
                        word_freq_uno.append(sheet.cell(cr,cc+1).value)
                        print(word_l[i])
                        word_lgood.append(sheet.cell(cr,cc+cck).value)
                        i += 1
                        cr = 1
                        break
            print('med')
            word_char.append(word_lgood)
            word_freq.append(word_freq_uno)
            #print (cr)
            #print(cc)
            #print (word_ldub,'word_ldub')
            word_l = word_ldub
            #print (word_l,'word_l')
            word_lgood = []
            word_freq_uno = []
            i = 0
            cck += 1
        
        print (word_char)
        print (word_freq)
        #print (word_char[1],'xi')
        #print (max(word_char[1]),'max(word_l)')
        #print (len(word_char), 'len_word_char')
        i = 0
        x = 0

        i = 0
        
        while i < len(word_freq):
            x = 0
            PmaxPi = []
            while x < len(word_freq[i]):
                try:
                    Pi = max(word_freq[i])/word_freq[i][x]
                except ArithmeticError:
                    Pi = 0
                
                PmaxPi.append(round(Pi,2))
                x += 1
                
            print (PmaxPi, 'PmaxPi')
            ki = PmaxPi
            ki[0] = ki[0]*4
            print (ki, 'ki')
            import numpy as np
            xiki = np.array(ki)*np.array(word_char[i])
            print (xiki)
            St = sum(xiki)/sum(ki)
            #print (round(St, 3))
            if St < 2.5:
                print ('слово подходит под данный критерий')
            else:
                print ('слово не подходит под данный критерий табл.')
            St_list.append(round(St, 3))
            i += 1
        St_list_all.append(St_list)
        print (St_list_all,'St_list_all')
        St_list = []
        r += 1


    h = 0
    print (len(St_list_all))
    St_coll_all = []
    while h < 25:
        St_coll = []
        St_sum = 0
        sla = 0
        while sla < len(St_list_all):
            St_sum = round(St_sum + St_list_all[sla][h],3)
            sla += 1
        St_avrg = round(St_sum/len(text_w), 3)
        #print (St_avrg, 'St_avrg')
        if St_avrg <= 2.5:
            St_coll.append(St_avrg)
            St_coll.append((sheet.cell(1,3+h).value))
        else:
            St_coll.append(St_avrg)
            St_coll.append((sheet.cell(50,3+h).value))
        print (St_coll)
        St_coll_all.append(St_coll)
        h += 1

    #print (St_coll_all)
    collect_list = dict(St_coll_all)
    xyr = sorted(collect_list.items(), key=operator.itemgetter(0))
    print (xyr)
    print (xyr[1])
    xyr_top = []
    h = 0
    while h < 10:
        try:
            xyr_top.append(xyr[h])
            h += 1
        except IndexError:
            xyr_top.append('-')
            h += 1
    print (xyr_top)



    return render_template('form.html', final=len(text_tokens),
                           text1=len(text_tokens),
                           text2=len(text_tokens),
                           text3=St_avrg,
                           text4=xyr_top[0],
                           text5=xyr_top[1],
                           text6=xyr_top[2],
                           text7=xyr_top[3],
                           text8=xyr_top[4],
                           text9=xyr_top[5],
                           text10=xyr_top[6],
                           text11=xyr_top[7],
                           text12=xyr_top[8])


if __name__ == "__main__":
    app.run(debug=True, host="127.0.0.5", port=5002, threaded=True)
